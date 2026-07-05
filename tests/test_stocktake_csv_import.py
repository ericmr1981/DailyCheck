"""盘点 xlsx 解析 + CSV 导入路由测试。"""
import io
from datetime import datetime

from openpyxl import Workbook

from tests.conftest import _wh


def _login_admin(client):
    with client.session_transaction() as s:
        s["user_id"] = 1
        s["warehouse_id"] = 1


def _make_xlsx_bytes(rows: list[tuple]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(1, 1, "title")
    ws.cell(2, 1, "分类"); ws.cell(2, 2, "物料名称")
    ws.cell(2, 6, "现有库存")
    for r_idx, row in enumerate(rows, start=3):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_skips_blank_name_and_defaults_none_to_zero():
    from blueprints.stocktake import parse_stocktake_xlsx
    # 行 1 = title, 行 2 = header, 行 3 起为数据
    rows = [
        (1, "title"),
        (2, "header"),
        (3, ("X", "A", None, None, None, 5, None, None, None)),
        (4, (None, None, None, None, None, None, None, None, None)),  # 合计行
        (5, ("X", "B", None, None, None, None, None, None, None)),  # None qty → 0
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r_idx, row in rows:
        for c_idx, val in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, val)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    result = parse_stocktake_xlsx(buf)
    assert len(result["rows"]) == 2
    assert result["rows"][0]["name"] == "A"
    assert result["rows"][0]["quantity"] == 5
    assert result["rows"][1]["name"] == "B"
    assert result["rows"][1]["quantity"] == 0


def test_csv_form_renders_for_admin(logged_client):
    client, _ = logged_client
    _login_admin(client)
    resp = client.get("/admin/stocktake-import")
    assert resp.status_code == 200


def test_csv_parse_requires_pending_batch(logged_client):
    """没有 pending batch 时,parse 应 flash 错误并回 /stocktake。"""
    client, wh_path = logged_client
    _login_admin(client)
    xlsx_bytes = _make_xlsx_bytes([("X", "AAA", None, None, None, 5, None, None, None)])
    data = {"file": (io.BytesIO(xlsx_bytes), "test.xlsx")}
    resp = client.post("/admin/stocktake-import", data=data,
                       content_type="multipart/form-data", follow_redirects=False)
    # 应重定向到 /stocktake (flash 错误)
    assert resp.status_code == 302
    assert resp.headers["Location"].endswith("/stocktake")


def test_csv_full_flow_writes_pending_batch(logged_client):
    """完整流程:start 创建 batch → 上传匹配 → commit 写入 stocktakes。"""
    client, wh_path = logged_client
    _login_admin(client)
    # 先 seed 一个品项
    conn = _wh(wh_path)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cur = conn.execute(
        "INSERT INTO categories (name, description, created_at) VALUES ('X', '', ?)",
        (ts,))
    cat_id = cur.lastrowid
    cur = conn.execute(
        "INSERT INTO items (sku, name, category_id, quantity, safety_stock, "
        "unit_cost, unit, gram_per_unit, updated_at) "
        "VALUES ('S1', '面粉', ?, 100, 0, 1, '件', 0, ?)",
        (cat_id, ts))
    item_id = cur.lastrowid
    conn.commit()
    conn.close()

    # 1. start 创建空 batch
    resp = client.post("/stocktake/start", follow_redirects=False)
    assert resp.status_code == 302
    # 验证 batch 已创建
    conn = _wh(wh_path)
    batch_id = conn.execute(
        "SELECT id FROM stocktake_batches WHERE status='pending' ORDER BY id DESC LIMIT 1"
    ).fetchone()["id"]
    conn.close()

    # 2. 上传 xlsx
    xlsx_bytes = _make_xlsx_bytes([
        ("X", "面粉", None, None, None, 95, None, None, None),  # diff = -5
    ])
    data = {"file": (io.BytesIO(xlsx_bytes), "test.xlsx")}
    resp = client.post("/admin/stocktake-import", data=data,
                       content_type="multipart/form-data", follow_redirects=False)
    assert resp.status_code == 302
    assert "/admin/stocktake-import/preview" in resp.headers["Location"]

    # 3. 预览页
    resp = client.get("/admin/stocktake-import/preview")
    assert resp.status_code == 200
    with client.session_transaction() as s:
        pv = s["stocktake_csv_preview"]
    assert len(pv["matched"]) == 1, f"expected 1 matched, got: {pv}"
    assert pv["matched"][0]["name"] == "面粉"

    # 4. commit
    resp = client.post("/admin/stocktake-import/commit", follow_redirects=False)
    assert resp.status_code == 302

    # 验证:stocktakes 行已写入, diff = -5
    conn = _wh(wh_path)
    conn.row_factory = __import__("sqlite3").Row
    st = conn.execute(
        "SELECT item_id, previous_quantity, actual_quantity, diff, batch_id "
        "FROM stocktakes WHERE batch_id=?", (batch_id,)
    ).fetchall()
    assert len(st) == 1
    assert st[0]["item_id"] == item_id
    assert st[0]["previous_quantity"] == 100
    assert st[0]["actual_quantity"] == 95
    assert st[0]["diff"] == -5
    conn.close()


def test_csv_skips_unknown_items(logged_client):
    """xlsx 中的品项系统里没有 → 计入 not_found,导入时跳过。"""
    client, wh_path = logged_client
    _login_admin(client)
    conn = _wh(wh_path)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cur = conn.execute(
        "INSERT INTO categories (name, description, created_at) VALUES ('X', '', ?)",
        (ts,))
    cat_id = cur.lastrowid
    conn.execute(
        "INSERT INTO items (sku, name, category_id, quantity, safety_stock, "
        "unit_cost, unit, gram_per_unit, updated_at) "
        "VALUES ('S1', '面粉', ?, 100, 0, 1, '件', 0, ?)",
        (cat_id, ts))
    conn.commit()
    conn.close()

    client.post("/stocktake/start", follow_redirects=False)

    xlsx_bytes = _make_xlsx_bytes([
        ("X", "面粉", None, None, None, 95, None, None, None),
        ("X", "不存在的物料", None, None, None, 10, None, None, None),
    ])
    data = {"file": (io.BytesIO(xlsx_bytes), "test.xlsx")}
    client.post("/admin/stocktake-import", data=data,
                content_type="multipart/form-data", follow_redirects=False)

    # 预览:matched=1, not_found=1
    with client.session_transaction() as s:
        pv = s["stocktake_csv_preview"]
    assert len(pv["matched"]) == 1
    assert len(pv["not_found"]) == 1
    assert pv["not_found"][0]["name"] == "不存在的物料"

    # commit:只写入 1 行
    client.post("/admin/stocktake-import/commit", follow_redirects=False)
    conn = _wh(wh_path)
    n = conn.execute("SELECT COUNT(*) AS c FROM stocktakes").fetchone()["c"]
    assert n == 1
    conn.close()


def test_csv_form_requires_admin(logged_client):
    """非 admin 应被拒。"""
    client, _ = logged_client
    # 不调 _login_admin,但 logged_client 默认已登录 user_id=1
    # 实际测试:用未登录 client
    client2 = client.application.test_client()
    resp = client2.get("/admin/stocktake-import")
    assert resp.status_code in (302, 403)