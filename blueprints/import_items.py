"""Web 端品项批量导入:上传 → 预览 → 确认写入。"""
from __future__ import annotations

import sqlite3
from contextlib import closing
from pathlib import Path
from typing import IO

from flask import (
    Blueprint, abort, current_app, flash, redirect, render_template, request, session, url_for,
)

from db import get_master_db
from permissions import require_login, require_role
from werkzeug.datastructures import FileStorage


def parse_xlsx(file_stream: IO[bytes]) -> dict:
    """解析 xlsx 文件为分组预览数据。

    期望格式:Sheet1,跳过前 2 行(标题 + 表头),第 3 行起为数据。
    分类列首行有值,后续空 → 沿用上一行(向下合并)。
    列序:分类(1)|物料名称(2)|SKU(3)|规格(4)|单价/元(5)|
          现有库存(6)|单位(7)|隐藏栏/盘点单位单价(8)|库存金额(9)

    Returns:
        {
            "groups_order": ["常温物料", ...],
            "groups_rows": {"常温物料": [{"row": 3, "name": "...", "spec": "...",
                                          "unit_cost": float, "unit": "str"}, ...]},
            "problems": ["行 N: 缺少分类", ...]
        }

    注意:本函数不写库,不读 session,纯函数。
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_stream, data_only=True)
    ws = wb["Sheet1"]
    groups_order: list[str] = []
    groups_rows: dict[str, list[dict]] = {}
    problems: list[str] = []
    prev_cat: str | None = None

    for r in range(3, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if not name:
            # 合计行 / 空行 → 跳过
            continue
        cat = ws.cell(r, 1).value or prev_cat
        if cat is None:
            problems.append(f"行 {r}: 缺少分类")
            continue
        prev_cat = cat
        if cat not in groups_rows:
            groups_order.append(cat)
            groups_rows[cat] = []
        unit_cost_raw = ws.cell(r, 8).value
        try:
            unit_cost = float(unit_cost_raw) if unit_cost_raw is not None else 0.0
        except (TypeError, ValueError):
            unit_cost = 0.0
            problems.append(f"行 {r}: 单价无法解析为数字 → 记为 0")
        unit_raw = ws.cell(r, 7).value
        unit = str(unit_raw).strip() if unit_raw is not None else ""
        if not unit:
            unit = "件"
        spec_raw = ws.cell(r, 4).value
        groups_rows[cat].append({
            "row": r,
            "name": str(name).strip(),
            "spec": str(spec_raw) if spec_raw is not None else None,
            "unit_cost": unit_cost,
            "unit": unit,
        })

    return {
        "groups_order": groups_order,
        "groups_rows": groups_rows,
        "problems": problems,
    }


bp = Blueprint("import_items", __name__, url_prefix="/admin/import-items")


@bp.route("", methods=["GET"])
@require_login
@require_role("admin")
def upload_form():
    """渲染上传表单。"""
    master = get_master_db()
    warehouses = master.execute(
        "SELECT code, name FROM warehouses ORDER BY id"
    ).fetchall()
    return render_template("admin/import_items.html", warehouses=warehouses)


@bp.route("", methods=["POST"])
@require_login
@require_role("admin")
def upload_parse():
    """解析上传的 xlsx → 缓存到 session → 重定向预览。"""
    file = request.files.get("file")
    warehouse_code = request.form.get("warehouse_code", "").strip()

    if not isinstance(file, FileStorage) or not file.filename:
        flash("请选择文件")
        return redirect(url_for("import_items.upload_form"))

    if not file.filename.lower().endswith(".xlsx"):
        flash("仅支持 .xlsx 文件")
        return redirect(url_for("import_items.upload_form"))

    # 校验仓库存在
    master = get_master_db()
    row = master.execute(
        "SELECT 1 FROM warehouses WHERE code = ?", (warehouse_code,)
    ).fetchone()
    if row is None:
        flash(f"仓库 {warehouse_code} 不存在")
        return redirect(url_for("import_items.upload_form"))

    try:
        preview = parse_xlsx(file.stream)
    except Exception:
        current_app.logger.exception("Excel 解析失败")
        flash("Excel 解析失败,请检查文件格式")
        return redirect(url_for("import_items.upload_form"))

    if not preview["groups_order"]:
        flash("Excel 中无数据行")
        return redirect(url_for("import_items.upload_form"))

    session["import_preview"] = {
        "warehouse_code": warehouse_code,
        "filename": file.filename,
        **preview,
    }
    return redirect(url_for("import_items.preview"))


@bp.route("/preview", methods=["GET"])
@require_login
@require_role("admin")
def preview():
    """渲染预览页。"""
    pv = session.get("import_preview")
    if not pv:
        flash("预览已过期,请重新上传")
        return redirect(url_for("import_items.upload_form"))

    master = get_master_db()
    wh = master.execute(
        "SELECT name FROM warehouses WHERE code = ?", (pv["warehouse_code"],)
    ).fetchone()
    target_wh_name = wh["name"] if wh else pv["warehouse_code"]

    total_rows = sum(len(v) for v in pv["groups_rows"].values())
    return render_template(
        "admin/import_items_preview.html",
        target_wh_name=target_wh_name,
        total_rows=total_rows,
        **pv,
    )


@bp.route("/commit", methods=["POST"])
@require_login
@require_role("admin")
def commit():
    """事务化:DELETE 预览分组下 items + INSERT 预览数据,缺品类则拒绝。"""
    from blueprints._helpers import gen_sku, now
    from config import BASE_DIR

    pv = session.pop("import_preview", None)
    if not pv:
        flash("预览已过期,请重新上传")
        return redirect(url_for("import_items.upload_form"))

    warehouse_code = pv["warehouse_code"]
    groups_order = pv["groups_order"]

    # 1. 校验仓库存在
    master = get_master_db()
    wh_row = master.execute(
        "SELECT db_path, name FROM warehouses WHERE code = ?", (warehouse_code,)
    ).fetchone()
    if wh_row is None:
        flash(f"仓库 {warehouse_code} 不存在")
        return redirect(url_for("import_items.upload_form"))

    db_path = Path(BASE_DIR) / wh_row["db_path"]

    # 3. 事务化:创建缺失分类 + DELETE 子表 + DELETE items + INSERT
    inserted = 0
    created_cats: list[str] = []
    try:
        with closing(sqlite3.connect(db_path)) as conn:
            conn.row_factory = sqlite3.Row
            conn.execute("PRAGMA foreign_keys = ON")
            ts = now()
            # 3a. 自动创建缺失的分组分类(同事务,保证原子性)
            existing_cats = {
                r["name"] for r in conn.execute("SELECT name FROM categories").fetchall()
            }
            for cat_name in groups_order:
                if cat_name not in existing_cats:
                    conn.execute(
                        "INSERT INTO categories (name, description, created_at) "
                        "VALUES (?, ?, ?)",
                        (cat_name, "导入自动创建", ts),
                    )
                    created_cats.append(cat_name)
            placeholders = ",".join("?" for _ in groups_order)
            # 子查询:本次导入分组下的所有 item_id
            items_subq = (
                f"item_id IN (SELECT id FROM items WHERE category_id IN "
                f"(SELECT id FROM categories WHERE name IN ({placeholders})))"
            )
            # 3b. 先清掉引用这些 item 的子行,避免 FK 约束阻止 DELETE items
            for table in [
                "stock_movements",
                "stocktakes",
                "restock_requests",
                "outbound_requests",
                "adjustment_requests",
                "product_bom",
                "production_run_items",
            ]:
                conn.execute(f"DELETE FROM {table} WHERE {items_subq}", groups_order)
            # 3c. 现在可以安全删 items
            conn.execute(
                f"DELETE FROM items WHERE category_id IN "
                f"(SELECT id FROM categories WHERE name IN ({placeholders}))",
                groups_order,
            )
            cat_by_name = {
                r["name"]: r["id"]
                for r in conn.execute("SELECT id, name FROM categories").fetchall()
            }
            for cat_name in groups_order:
                cid = cat_by_name[cat_name]
                for item in pv["groups_rows"][cat_name]:
                    conn.execute(
                        """INSERT INTO items
                           (sku, name, category_id, quantity, safety_stock,
                            unit_cost, unit, gram_per_unit, updated_at)
                           VALUES (?, ?, ?, 0, 0, ?, ?, 0, ?)""",
                        (gen_sku(), item["name"], cid,
                         item["unit_cost"], item["unit"], ts),
                    )
                    inserted += 1
            conn.commit()
    except sqlite3.IntegrityError:
        flash("导入失败,请重试")
        return redirect(url_for("import_items.upload_form"))

    # 4. audit
    from blueprints.auth import audit
    audit("import_items.import", "warehouse", warehouse_code, {
        "count": inserted,
        "filename": pv.get("filename"),
        "created_categories": created_cats,
    })
    msg = f"导入成功:{inserted} 条品项"
    if created_cats:
        msg += f"(自动创建分类:{', '.join(created_cats)})"
    flash(msg)
    return redirect(url_for("items.items_list"))
