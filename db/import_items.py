"""One-shot import: read 产品模版.xlsx, insert items into a target warehouse db.

Usage:
    python3 -m db.import_items /Users/ericmr/Downloads/产品模版.xlsx wh_002

Required columns (in this order, with header row 1):
    品项编码 | 品项名称 | 品项类别 | 订货单位 | 单价

Per-item defaults: quantity=0, safety_stock=0.
unit_cost comes from the xlsx 单价 column (renamed from "成本价" elsewhere).
SKU is set to the 品项编码 (e.g. "WP0190") — empty / None SKU falls back
to the existing AUTO- timestamp generator.

This is idempotent on category: it never duplicates a category row. It does
NOT deduplicate items: re-running inserts duplicate items. Use with care.
"""
from __future__ import annotations

import sqlite3
import sys
from contextlib import closing
from datetime import datetime
from pathlib import Path

from config import BASE_DIR


def now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def import_items(xlsx_path: str, warehouse_code: str) -> None:
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl is required: pip install openpyxl")

    # Locate the warehouse db via master.db.
    master = sqlite3.connect(BASE_DIR / "db" / "master.db")
    master.row_factory = sqlite3.Row
    wh = master.execute(
        "SELECT db_path FROM warehouses WHERE code = ?", (warehouse_code,)
    ).fetchone()
    if wh is None:
        sys.exit(f"Warehouse '{warehouse_code}' not found in master.db")
    db_path = BASE_DIR / wh["db_path"]
    master.close()

    if not db_path.exists():
        sys.exit(f"Warehouse db file missing: {db_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # Read header → column index map
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    required = ["品项编码", "品项名称", "品项类别", "订货单位", "单价"]
    try:
        col = {h: headers.index(h) + 1 for h in required}
    except ValueError as e:
        sys.exit(f"Missing required column: {e}. Headers were: {headers}")

    with closing(sqlite3.connect(db_path)) as conn:
        conn.row_factory = sqlite3.Row
        # Build category id cache from categories table
        cat_rows = conn.execute("SELECT id, name FROM categories").fetchall()
        cat_by_name = {r["name"]: r["id"] for r in cat_rows}

        # Detect unknown categories before inserting anything
        unknown: set[str] = set()
        data_rows = []
        for r in range(2, ws.max_row + 1):
            sku = ws.cell(r, col["品项编码"]).value
            name = ws.cell(r, col["品项名称"]).value
            cat = ws.cell(r, col["品项类别"]).value
            unit = ws.cell(r, col["订货单位"]).value or "件"
            price = ws.cell(r, col["单价"]).value
            if not name:
                continue
            data_rows.append((sku, name, cat, unit, price))
            if cat and cat not in cat_by_name:
                unknown.add(cat)

        if unknown:
            sys.exit(
                "Unknown categories in xlsx (not in target warehouse's categories, "
                "refusing to insert): " + ", ".join(sorted(unknown))
            )

        inserted = 0
        skipped = 0
        for sku, name, cat, unit, price in data_rows:
            # Skip blank-SKU inserts when an item with the same name already exists
            existing = conn.execute(
                "SELECT 1 FROM items WHERE name = ?", (name,)
            ).fetchone()
            if existing is not None:
                skipped += 1
                continue
            sku_value = sku if sku else f"AUTO-{datetime.now().strftime('%Y%m%d%H%M%S%f')}"
            try:
                unit_cost = float(price) if price is not None else 0.0
            except (TypeError, ValueError):
                unit_cost = 0.0
            conn.execute(
                """INSERT INTO items
                   (sku, name, category_id, quantity, safety_stock, unit_cost, unit, updated_at)
                   VALUES (?, ?, ?, 0, 0, ?, ?, ?)""",
                (sku_value, name, cat_by_name[cat], unit_cost, unit, now()),
            )
            inserted += 1
        conn.commit()

    print(f"Warehouse: {warehouse_code} ({db_path})")
    print(f"Inserted:  {inserted}")
    print(f"Skipped:   {skipped} (name already exists)")
    print(f"Total rows in xlsx: {len(data_rows)}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        sys.exit("Usage: python3 -m db.import_items <xlsx_path> <warehouse_code>")
    import_items(sys.argv[1], sys.argv[2])
