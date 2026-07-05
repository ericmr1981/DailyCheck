"""上传 + 预览 + commit 路由测试。"""
import io
import sqlite3
from datetime import datetime

from openpyxl import Workbook


def _login_admin(client):
    """直接设 session 为 admin(role='admin' on warehouse)。"""
    with client.session_transaction() as s:
        s["user_id"] = 1
        s["warehouse_id"] = 1


def _make_xlsx_bytes(rows: list[tuple]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(1, 1, "title")
    ws.cell(2, 1, "分类"); ws.cell(2, 2, "物料名称")
    ws.cell(2, 7, "单位"); ws.cell(2, 8, "隐藏栏/盘点单位单价")
    for r_idx, row in enumerate(rows, start=3):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_upload_form_renders_for_admin(logged_client):
    client, _ = logged_client
    _login_admin(client)
    resp = client.get("/admin/import-items")
    assert resp.status_code == 200
    # 检查页面中至少包含一个核心标识词
    assert b"\xe6\x9c\x8d\xe5\x8a\xa1\xe5\x99\xa8\xe7\x9b\xae\xe5\x89\x8d" in resp.data or \
           b"\xe5\x93\x81\xe9\xa1\xb9\xe6\x89\xb9\xe9\x87\x8f\xe5\xaf\xbc\xe5\x85\xa5" in resp.data


def test_upload_parse_redirects_to_preview(logged_client):
    client, _ = logged_client
    _login_admin(client)
    xlsx_bytes = _make_xlsx_bytes([
        ("X", "A", None, "s", 100, 5, "箱", 10, 50),
    ])
    data = {
        "file": (io.BytesIO(xlsx_bytes), "test.xlsx"),
        "warehouse_code": "wh_test",
    }
    resp = client.post("/admin/import-items", data=data,
                       content_type="multipart/form-data", follow_redirects=False)
    assert resp.status_code == 302
    assert "/admin/import-items/preview" in resp.headers["Location"]


def test_upload_rejects_non_xlsx(logged_client):
    client, _ = logged_client
    _login_admin(client)
    data = {
        "file": (io.BytesIO(b"not an xlsx"), "test.txt"),
        "warehouse_code": "wh_test",
    }
    resp = client.post("/admin/import-items", data=data,
                       content_type="multipart/form-data", follow_redirects=False)
    # 拒绝: 重定向回 form
    assert resp.status_code == 302
    assert resp.headers["Location"] == "/admin/import-items"


def test_preview_without_session_redirects(logged_client):
    client, _ = logged_client
    _login_admin(client)
    resp = client.get("/admin/import-items/preview", follow_redirects=False)
    # 没有 session 缓存 → 重定向回 form
    assert resp.status_code == 302
    assert resp.headers["Location"] == "/admin/import-items"


def test_unauthenticated_cannot_access(logged_client):
    """未登录访问 /admin/import-items 应被拒。"""
    client, _wh_path = logged_client
    # 用全新 client,没有 session → 应重定向登录或 403
    client2 = client.application.test_client()
    resp = client2.get("/admin/import-items")
    assert resp.status_code in (302, 403)


# ---------------------------------------------------------------------------
# commit 路由测试 (Task 4)
# ---------------------------------------------------------------------------


def test_commit_requires_session(logged_client):
    """无 session 缓存 → commit 拒绝,items 表不应有新增。"""
    client, wh_path = logged_client
    _login_admin(client)
    resp = client.post("/admin/import-items/commit", follow_redirects=True)
    assert resp.status_code in (200, 400)
    wh = sqlite3.connect(wh_path)
    wh.row_factory = sqlite3.Row
    n = wh.execute("SELECT COUNT(*) AS c FROM items").fetchone()["c"]
    assert n == 0
    wh.close()


def test_commit_creates_missing_categories(logged_client):
    """目标仓库缺 xlsx 的某个分组 → commit 自动创建分类,不拒绝。"""
    client, wh_path = logged_client
    _login_admin(client)
    xlsx_bytes = _make_xlsx_bytes([
        ("不存在的分类", "A", None, "s", 100, 5, "箱", 10, 50),
    ])
    data = {"file": (io.BytesIO(xlsx_bytes), "test.xlsx"),
            "warehouse_code": "wh_test"}
    client.post("/admin/import-items", data=data,
                content_type="multipart/form-data", follow_redirects=False)
    resp = client.post("/admin/import-items/commit", follow_redirects=False)
    assert resp.status_code == 302  # 重定向到 /items
    wh = sqlite3.connect(wh_path)
    wh.row_factory = sqlite3.Row
    # 验证:新分类被自动创建
    cat = wh.execute(
        "SELECT name, description FROM categories WHERE name=?",
        ("不存在的分类",),
    ).fetchone()
    assert cat is not None
    assert cat["description"] == "导入自动创建"
    # 验证:item 被插入
    n = wh.execute("SELECT COUNT(*) AS c FROM items").fetchone()["c"]
    assert n == 1
    wh.close()


def test_commit_inserts_items_when_categories_match(logged_client):
    """仓库已有"X"分类 + 上传"X"分类 → commit 后 items 行数 = 2。"""
    client, wh_path = logged_client
    _login_admin(client)
    # 注入一个 "X" 分类
    wh = sqlite3.connect(wh_path)
    wh.row_factory = sqlite3.Row
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    wh.execute(
        "INSERT INTO categories (name, description, created_at) VALUES ('X', '', ?)",
        (ts,))
    wh.commit()
    wh.close()
    xlsx_bytes = _make_xlsx_bytes([
        ("X", "AAA", None, "spec1", 100, 5, "箱", 10, 50),
        (None, "BBB", None, "spec2", 200, 3, "包", 20, 60),
    ])
    data = {"file": (io.BytesIO(xlsx_bytes), "test.xlsx"),
            "warehouse_code": "wh_test"}
    client.post("/admin/import-items", data=data,
                content_type="multipart/form-data", follow_redirects=False)
    resp = client.post("/admin/import-items/commit", follow_redirects=False)
    assert resp.status_code == 302
    wh = sqlite3.connect(wh_path)
    wh.row_factory = sqlite3.Row
    rows = wh.execute(
        "SELECT name, unit, unit_cost FROM items ORDER BY id"
    ).fetchall()
    assert len(rows) == 2
    assert rows[0]["name"] == "AAA"
    assert rows[0]["unit"] == "箱"
    assert rows[0]["unit_cost"] == 10.0
    assert rows[1]["unit"] == "包"
    wh.close()


def test_commit_is_idempotent(logged_client):
    """二次 commit → 行数仍为 1(覆盖无残留)。"""
    client, wh_path = logged_client
    _login_admin(client)
    wh = sqlite3.connect(wh_path)
    wh.row_factory = sqlite3.Row
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    wh.execute(
        "INSERT INTO categories (name, description, created_at) VALUES ('X', '', ?)",
        (ts,))
    wh.commit()
    wh.close()
    xlsx_bytes = _make_xlsx_bytes([
        ("X", "AAA", None, "s", 100, 5, "箱", 10, 50),
    ])

    def _post_upload():
        return client.post(
            "/admin/import-items",
            data={"file": (io.BytesIO(xlsx_bytes), "test.xlsx"),
                  "warehouse_code": "wh_test"},
            content_type="multipart/form-data",
            follow_redirects=False,
        )

    _post_upload()
    client.post("/admin/import-items/commit", follow_redirects=False)
    # 再来一遍(用全新 BytesIO,因为上次的已被消费)
    _post_upload()
    client.post("/admin/import-items/commit", follow_redirects=False)
    wh = sqlite3.connect(wh_path)
    wh.row_factory = sqlite3.Row
    n = wh.execute("SELECT COUNT(*) AS c FROM items").fetchone()["c"]
    assert n == 1
    wh.close()


def test_commit_purges_child_rows_when_deleting_items(logged_client):
    """commit() 必须先 DELETE 子行(stock_movements 等)再删 items,
    否则带历史仓库无法重新导入。"""
    client, wh_path = logged_client
    _login_admin(client)
    wh = sqlite3.connect(wh_path)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cur = wh.execute(
        "INSERT INTO categories (name, description, created_at) VALUES ('X', '', ?)",
        (ts,))
    cat_id = cur.lastrowid
    cur = wh.execute(
        "INSERT INTO items (sku, name, category_id, quantity, safety_stock, "
        "unit_cost, unit, gram_per_unit, updated_at) "
        "VALUES ('OLD', 'OldItem', ?, 0, 0, 1, '件', 0, ?)",
        (cat_id, ts))
    item_id = cur.lastrowid
    wh.execute(
        "INSERT INTO stock_movements (item_id, action, delta, created_at) "
        "VALUES (?, 'restock', 10, ?)",
        (item_id, ts))
    wh.commit()
    wh.close()
    xlsx_bytes = _make_xlsx_bytes([("X", "NewItem", None, "s", 100, 5, "箱", 10, 50)])
    data = {"file": (io.BytesIO(xlsx_bytes), "test.xlsx"),
            "warehouse_code": "wh_test"}
    client.post("/admin/import-items", data=data,
                content_type="multipart/form-data", follow_redirects=False)
    resp = client.post("/admin/import-items/commit", follow_redirects=False)
    assert resp.status_code == 302
    wh = sqlite3.connect(wh_path)
    wh.row_factory = sqlite3.Row
    names = [r["name"] for r in wh.execute("SELECT name FROM items").fetchall()]
    assert names == ["NewItem"]
    n_mov = wh.execute("SELECT COUNT(*) AS c FROM stock_movements").fetchone()["c"]
    assert n_mov == 0
    wh.close()
