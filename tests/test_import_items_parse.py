"""解析 泰柯盘点表.xlsx 单元测试。"""
import io
from pathlib import Path

import pytest
from openpyxl import Workbook


def _make_xlsx(path: Path, rows: list[tuple]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(1, 1, "tamkoko月盘表 ")
    ws.cell(2, 1, "分类")
    ws.cell(2, 2, "物料名称")
    ws.cell(2, 3, "SKU")
    ws.cell(2, 4, "规格")
    ws.cell(2, 5, "单价/元")
    ws.cell(2, 6, "现有库存")
    ws.cell(2, 7, "单位")
    ws.cell(2, 8, "隐藏栏/盘点单位单价")
    ws.cell(2, 9, "库存金额")
    for r_idx, row in enumerate(rows, start=3):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(r_idx, c_idx, val)
    wb.save(str(path))


def test_parse_groups_by_category(tmp_path):
    """5 分组向下合并:第一行有分类名,后续行空 → 沿用上一行。"""
    rows = [
        ("常温物料", "A", None, "x", 100, 5, "箱", 10, 50),
        (None, "B", None, "y", 200, 3, "包", 20, 60),
        ("包装材料", "C", None, "z", 300, 2, "个", 30, 60),
    ]
    p = tmp_path / "in.xlsx"
    _make_xlsx(p, rows)
    from blueprints.import_items import parse_xlsx
    with open(p, "rb") as f:
        result = parse_xlsx(f)
    assert result["groups_order"] == ["常温物料", "包装材料"]
    assert len(result["groups_rows"]["常温物料"]) == 2
    assert result["groups_rows"]["常温物料"][0]["name"] == "A"
    assert result["groups_rows"]["常温物料"][0]["unit_cost"] == 10.0
    assert result["groups_rows"]["常温物料"][1]["unit"] == "包"


def test_parse_none_unit_cost_defaults_to_zero(tmp_path):
    rows = [("X", "A", None, "x", 100, None, "箱", None, 0)]
    p = tmp_path / "in.xlsx"
    _make_xlsx(p, rows)
    from blueprints.import_items import parse_xlsx
    with open(p, "rb") as f:
        result = parse_xlsx(f)
    assert result["groups_rows"]["X"][0]["unit_cost"] == 0.0


def test_parse_empty_unit_defaults_to_jian(tmp_path):
    rows = [("X", "A", None, "x", 100, 5, None, 10, 50)]
    p = tmp_path / "in.xlsx"
    _make_xlsx(p, rows)
    from blueprints.import_items import parse_xlsx
    with open(p, "rb") as f:
        result = parse_xlsx(f)
    assert result["groups_rows"]["X"][0]["unit"] == "件"


def test_parse_skips_blank_name_rows(tmp_path):
    """合计行(名称为空)跳过。"""
    rows = [
        ("X", "A", None, "x", 100, 5, "箱", 10, 50),
        (None, None, None, None, None, None, "总额", None, 13888.106),
    ]
    p = tmp_path / "in.xlsx"
    _make_xlsx(p, rows)
    from blueprints.import_items import parse_xlsx
    with open(p, "rb") as f:
        result = parse_xlsx(f)
    assert len(result["groups_rows"]["X"]) == 1


def test_parse_records_problem_when_first_row_missing_category(tmp_path):
    """首行分类为空 → 计入 problems,跳过该行。"""
    rows = [(None, "A", None, "x", 100, 5, "箱", 10, 50)]
    p = tmp_path / "in.xlsx"
    _make_xlsx(p, rows)
    from blueprints.import_items import parse_xlsx
    with open(p, "rb") as f:
        result = parse_xlsx(f)
    assert result["groups_rows"] == {}
    assert any("缺少分类" in p for p in result["problems"])
