"""Stocktake workflows: start a batch, fill in actuals, submit, rollback,
edit, approve, CSV import. Approval requires manager role.
"""
from __future__ import annotations

import io
from typing import IO

from flask import (
    Blueprint, current_app, flash, g, redirect, render_template, request,
    session, url_for,
)
from werkzeug.datastructures import FileStorage

from db import get_warehouse_db
from permissions import require_login, require_role
from ._helpers import now, parse_qty
from .auth import audit


bp = Blueprint("stocktake", __name__)


# ---------------------------------------------------------------------------
# xlsx 解析(纯函数):泰柯盘点表格式 → 物料盘点清单
# ---------------------------------------------------------------------------

def parse_stocktake_xlsx(file_stream: IO[bytes]) -> dict:
    """解析 xlsx 为盘点预览数据。

    期望格式(泰柯盘点表):
      Sheet1, 跳过前 2 行(标题 + 表头), 第 3 行起为数据。
      列 2 = 物料名称(必填)
      列 6 = 现有库存(实际盘点数量, None → 0)
      其他列忽略(分类/SKU/规格/单价/单位/盘点单价/库存金额)。

    Returns:
        {
            "rows": [{"row": int, "name": str, "quantity": float}, ...],
            "problems": ["行 N: ..."],
        }
    """
    import openpyxl
    wb = openpyxl.load_workbook(file_stream, data_only=True)
    ws = wb["Sheet1"]
    rows: list[dict] = []
    problems: list[str] = []
    for r in range(3, ws.max_row + 1):
        name_raw = ws.cell(r, 2).value
        if not name_raw:
            continue  # 合计行 / 空行
        name = str(name_raw).strip()
        qty_raw = ws.cell(r, 6).value
        try:
            quantity = float(qty_raw) if qty_raw is not None else 0.0
        except (TypeError, ValueError):
            quantity = 0.0
            problems.append(f"行 {r}: {name} 盘点数量无法解析 → 记为 0")
        rows.append({"row": r, "name": name, "quantity": quantity})
    return {"rows": rows, "problems": problems}


# ---------------------------------------------------------------------------
# CSV 导入流程:上传 → 预览 → commit 写入最近 pending batch
# ---------------------------------------------------------------------------

@bp.route("/admin/stocktake-import", methods=["GET"])
@require_login
@require_role("admin")
def stocktake_csv_form():
    """渲染上传表单。目标仓库 = g.warehouse,目标 batch = 最近 pending batch。"""
    db = get_warehouse_db()
    pending = db.execute(
        "SELECT id FROM stocktake_batches WHERE status='pending' AND rolled_back=0 "
        "ORDER BY id DESC LIMIT 1"
    ).fetchone()
    return render_template(
        "admin/stocktake_import.html",
        pending_batch_id=pending["id"] if pending else None,
    )


@bp.route("/admin/stocktake-import", methods=["POST"])
@require_login
@require_role("admin")
def stocktake_csv_parse():
    """解析 xlsx → 与 pending batch 匹配 → 缓存预览。"""
    file = request.files.get("file")

    if not isinstance(file, FileStorage) or not file.filename:
        flash("请选择文件")
        return redirect(url_for("stocktake.stocktake_csv_form"))

    if not file.filename.lower().endswith(".xlsx"):
        flash("仅支持 .xlsx 文件")
        return redirect(url_for("stocktake.stocktake_csv_form"))

    try:
        preview = parse_stocktake_xlsx(file.stream)
    except Exception:
        current_app.logger.exception("盘点 xlsx 解析失败")
        flash("Excel 解析失败,请检查文件格式")
        return redirect(url_for("stocktake.stocktake_csv_form"))

    if not preview["rows"]:
        flash("Excel 中无数据行")
        return redirect(url_for("stocktake.stocktake_csv_form"))

    # 必须有 pending batch
    db = get_warehouse_db()
    pending = db.execute(
        "SELECT id FROM stocktake_batches WHERE status='pending' AND rolled_back=0 "
        "ORDER BY id DESC LIMIT 1"
    ).fetchone()
    if pending is None:
        flash("请先在盘点页面点击「开始盘点」创建批次")
        return redirect(url_for("stocktake.stocktake_list"))

    # 匹配品项:按 name 查 items
    items_rows = db.execute(
        "SELECT id, name, quantity FROM items"
    ).fetchall()
    by_name: dict[str, list] = {}
    for it in items_rows:
        by_name.setdefault(it["name"], []).append(it)

    matched: list[dict] = []
    not_found: list[dict] = []
    for row in preview["rows"]:
        candidates = by_name.get(row["name"])
        if not candidates:
            not_found.append({"row": row["row"], "name": row["name"], "quantity": row["quantity"]})
            continue
        if len(candidates) > 1:
            preview["problems"].append(
                f"行 {row['row']}: {row['name']} 在系统中有 {len(candidates)} 个同名品项,取第一个"
            )
        cand = candidates[0]
        matched.append({
            "row": row["row"],
            "item_id": cand["id"],
            "name": cand["name"],
            "previous_quantity": cand["quantity"],
            "actual_quantity": row["quantity"],
            "diff": round(row["quantity"] - cand["quantity"], 2),
        })

    session["stocktake_csv_preview"] = {
        "filename": file.filename,
        "batch_id": pending["id"],
        "matched": matched,
        "not_found": not_found,
        "problems": preview["problems"],
    }
    return redirect(url_for("stocktake.stocktake_csv_preview_view"))


@bp.route("/admin/stocktake-import/preview", methods=["GET"])
@require_login
@require_role("admin")
def stocktake_csv_preview_view():
    pv = session.get("stocktake_csv_preview")
    if not pv:
        flash("预览已过期,请重新上传")
        return redirect(url_for("stocktake.stocktake_csv_form"))
    return render_template("admin/stocktake_import_preview.html", **pv)


@bp.route("/admin/stocktake-import/commit", methods=["POST"])
@require_login
@require_role("admin")
def stocktake_csv_commit():
    pv = session.pop("stocktake_csv_preview", None)
    if not pv:
        flash("预览已过期,请重新上传")
        return redirect(url_for("stocktake.stocktake_csv_form"))

    batch_id = pv["batch_id"]
    matched = pv["matched"]
    if not matched:
        flash("无可写入的行(全部未匹配)")
        return redirect(url_for("stocktake.stocktake_csv_form"))

    db = get_warehouse_db()
    # 校验 batch 仍存在且 pending
    batch = db.execute(
        "SELECT id, status FROM stocktake_batches WHERE id=? AND rolled_back=0",
        (batch_id,),
    ).fetchone()
    if batch is None or batch["status"] != "pending":
        flash(f"批次 #{batch_id} 已不存在或非 pending 状态")
        return redirect(url_for("stocktake.stocktake_list"))

    ts = now()
    inserted = 0
    for m in matched:
        db.execute(
            """INSERT INTO stocktakes
               (item_id, previous_quantity, actual_quantity, diff, batch_id, created_at, note)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (m["item_id"], m["previous_quantity"], m["actual_quantity"],
             m["diff"], batch_id, ts, "CSV 导入"),
        )
        inserted += 1
    db.commit()
    audit("stocktake.csv_import", "batch", batch_id, {
        "count": inserted,
        "filename": pv["filename"],
        "not_found_count": len(pv["not_found"]),
    })
    msg = f"盘点导入成功:{inserted} 条已写入批次 #{batch_id}"
    if pv["not_found"]:
        msg += f"({len(pv['not_found'])} 条未匹配,见预览页)"
    flash(msg)
    return redirect(url_for("stocktake.stocktake_list"))


@bp.route("/stocktake", methods=["GET"])
@require_login
def stocktake_list():
    db = get_warehouse_db()
    batches = db.execute(
        """SELECT b.id, b.created_at, b.note, b.status, b.rolled_back, COUNT(s.id) AS item_count
           FROM stocktake_batches b
           LEFT JOIN stocktakes s ON s.batch_id = b.id
           GROUP BY b.id ORDER BY b.id DESC LIMIT 20"""
    ).fetchall()
    return render_template("stocktake.html", batches=batches)


@bp.route("/stocktake/start", methods=["POST"])
@require_login
def stocktake_start():
    """开始盘点 = 创建一个空 batch(status=pending, 无 stocktakes 行)。
    后续可通过 CSV 导入或 /stocktake/session 手动填写。
    已有 pending batch 时直接跳 session 页(避免重复创建空 batch)。"""
    db = get_warehouse_db()
    existing = db.execute(
        "SELECT id FROM stocktake_batches WHERE status='pending' AND rolled_back=0 "
        "ORDER BY id DESC LIMIT 1"
    ).fetchone()
    if existing is None:
        cur = db.execute(
            "INSERT INTO stocktake_batches (created_at, note, status, rolled_back) "
            "VALUES (?, '', 'pending', 0)",
            (now(),),
        )
        db.commit()
        audit("stocktake.start", "batch", cur.lastrowid, {})
    return redirect(url_for("stocktake.stocktake_session"))


@bp.route("/stocktake/session", methods=["GET"])
@require_login
def stocktake_session():
    db = get_warehouse_db()
    items_data = db.execute(
        """SELECT i.id, i.name, i.quantity, i.unit, i.safety_stock, c.name AS category_name
           FROM items i JOIN categories c ON c.id = i.category_id
           ORDER BY c.name, i.name"""
    ).fetchall()
    return render_template("stocktake_session.html", items=items_data)


@bp.route("/stocktake/submit", methods=["POST"])
@require_login
def stocktake_submit():
    db = get_warehouse_db()
    note = request.form.get("note", "").strip()
    items_data = db.execute("SELECT id, quantity FROM items").fetchall()

    changed_rows = []
    for item in items_data:
        raw = request.form.get(f"actual_{item['id']}", "").strip()
        if raw == "":
            continue
        actual_quantity = parse_qty(raw)
        previous_quantity = parse_qty(item["quantity"])
        diff = actual_quantity - previous_quantity
        changed_rows.append((int(item["id"]), previous_quantity, actual_quantity, diff))

    if not changed_rows:
        flash("请至少填写一个盘点数量")
        return redirect(url_for("stocktake.stocktake_session"))

    cur = db.execute(
        "INSERT INTO stocktake_batches (created_at, note, status, rolled_back) VALUES (?, ?, 'pending', 0)",
        (now(), note),
    )
    batch_id = cur.lastrowid
    for item_id, previous_quantity, actual_quantity, diff in changed_rows:
        db.execute(
            """INSERT INTO stocktakes
               (item_id, previous_quantity, actual_quantity, diff, batch_id, created_at, note)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (item_id, previous_quantity, actual_quantity, diff, batch_id, now(), note),
        )
    db.commit()
    audit("stocktake.submit", "batch", batch_id, {"count": len(changed_rows)})
    flash(f"盘点批次 #{batch_id} 已提交，等待审核")
    return redirect(url_for("stocktake.stocktake_list"))


@bp.route("/stocktake/batch/<int:batch_id>/rollback", methods=["POST"])
@require_role("staff")
def rollback(batch_id: int):
    db = get_warehouse_db()
    batch = db.execute(
        "SELECT id, status, rolled_back FROM stocktake_batches WHERE id = ?", (batch_id,)
    ).fetchone()
    if batch is None:
        flash("盘点批次不存在")
        return redirect(url_for("stocktake.stocktake_list"))
    if int(batch["rolled_back"]) == 1:
        flash("该批次已回滚，无需重复操作")
        return redirect(url_for("stocktake.stocktake_list"))

    records = db.execute(
        "SELECT item_id, diff FROM stocktakes WHERE batch_id = ?", (batch_id,)
    ).fetchall()
    for record in records:
        diff = parse_qty(record["diff"])
        if diff == 0:
            continue
        item_id = int(record["item_id"])
        db.execute(
            "UPDATE items SET quantity = quantity - ?, updated_at = ? WHERE id = ?",
            (diff, now(), item_id),
        )
        db.execute(
            """INSERT INTO stock_movements (item_id, action, delta, note, created_at)
               VALUES (?, '盘点回滚', ?, ?, ?)""",
            (item_id, -diff, f"回滚盘点批次#{batch_id}", now()),
        )

    # Reverse any synthetic outbound_requests written at approve time
    # so /summary no longer counts the loss as consumption. We hard-
    # delete the row rather than flipping rolled_back, because the
    # user's intent at this point is "undo the approval entirely".
    batch_row = db.execute(
        "SELECT loss_req_ids FROM stocktake_batches WHERE id=?", (batch_id,)
    ).fetchone()
    if batch_row and batch_row["loss_req_ids"]:
        for req_id in (
            int(x) for x in batch_row["loss_req_ids"].split(",") if x.strip()
        ):
            db.execute(
                "DELETE FROM outbound_requests WHERE id=? AND reason=?",
                (req_id, f"盘点审核#{batch_id}盘亏"),
            )

    db.execute(
        "UPDATE stocktake_batches SET rolled_back = 1, status = 'rolled_back' WHERE id = ?",
        (batch_id,),
    )
    db.commit()
    audit("stocktake.rollback", "batch", batch_id)
    flash(f"盘点批次 #{batch_id} 已回滚")
    return redirect(url_for("stocktake.stocktake_list"))


@bp.route("/stocktake/batch/<int:batch_id>/edit", methods=["GET"])
@require_role("staff")
def edit(batch_id: int):
    db = get_warehouse_db()
    batch = db.execute(
        "SELECT id, note, rolled_back FROM stocktake_batches WHERE id = ?", (batch_id,)
    ).fetchone()
    if batch is None:
        flash("盘点批次不存在")
        return redirect(url_for("stocktake.stocktake_list"))
    records = db.execute(
        """SELECT s.id, s.item_id, s.previous_quantity, s.actual_quantity, s.diff,
                  i.name AS item_name, i.quantity AS current_quantity, i.unit,
                  c.name AS category_name
           FROM stocktakes s
           JOIN items i ON i.id = s.item_id
           JOIN categories c ON c.id = i.category_id
           WHERE s.batch_id = ? ORDER BY c.name, i.name""",
        (batch_id,),
    ).fetchall()
    return render_template("stocktake_edit.html", batch=batch, records=records)


@bp.route("/stocktake/batch/<int:batch_id>/edit", methods=["POST"])
@require_role("staff")
def submit_edit(batch_id: int):
    db = get_warehouse_db()
    batch = db.execute(
        "SELECT id, rolled_back FROM stocktake_batches WHERE id = ?", (batch_id,)
    ).fetchone()
    if batch is None:
        flash("盘点批次不存在")
        return redirect(url_for("stocktake.stocktake_list"))
    if int(batch["rolled_back"]) == 1:
        flash("已回滚的批次不能修改")
        return redirect(url_for("stocktake.stocktake_list"))

    records = db.execute(
        "SELECT id, item_id, previous_quantity FROM stocktakes WHERE batch_id = ?",
        (batch_id,),
    ).fetchall()
    changed = 0
    for rec in records:
        raw = request.form.get(f"actual_{rec['item_id']}", "").strip()
        if raw == "":
            continue
        new_actual = parse_qty(raw)
        new_diff = new_actual - parse_qty(rec["previous_quantity"])
        if new_diff == 0:
            continue
        db.execute(
            "UPDATE stocktakes SET actual_quantity = ?, diff = ? WHERE id = ?",
            (new_actual, new_diff, int(rec["id"])),
        )
        db.execute(
            """INSERT INTO stock_movements (item_id, action, delta, note, created_at)
               VALUES (?, '盘点修正', ?, ?, ?)""",
            (int(rec["item_id"]), new_diff, f"修正盘点批次#{batch_id}", now()),
        )
        changed += 1
    if changed == 0:
        flash("未检测到变更")
    else:
        db.commit()
        audit("stocktake.edit", "batch", batch_id, {"changed": changed})
        flash(f"盘点批次 #{batch_id} 已更新 {changed} 项")
    return redirect(url_for("stocktake.stocktake_list"))


@bp.route("/stocktake/batch/<int:batch_id>/approve", methods=["POST"])
@require_role("staff")
def approve(batch_id: int):
    """Approval semantics (口径 B):
    - 盘亏 (diff < 0): real consumption. We write a synthetic
      outbound_requests row (rolled_back=0) so /summary counts it
      in total_consumed_value, AND the matching stock_movements
      action='出库' row for the flow / audit trail.
    - 盘盈 (diff > 0): only inventory adjustment (账面增, NOT counted
      as consumption). Written to stock_movements.action='库存调整'.
      Operators can use the adjustment-order page to write off a
      盘盈 that was actually a previous 出库-mis-entry.

    Idempotency: if the batch went through submit_edit first, every
    edited item carries a 盘点修正 movement whose note exactly reads
    `修正盘点批次#{batch_id}`. For those items we skip BOTH the items
    UPDATE AND the synthetic outbound_request / 库存调整 insertion —
    the edit path already recorded the corrected diff, so approving
    must not apply it again (double-deduction) nor emit a phantom
    consumption for /summary. The note is matched with `=` (not LIKE)
    so substring collisions across batch IDs (e.g. approving #9 vs
    a note referring to #99) don't cross-pollute the skip set.

    Rollback reverses both: deletes the synthetic outbound_requests,
    or flips rolled_back=1 if a later rollback is needed.
    """
    db = get_warehouse_db()
    batch = db.execute(
        "SELECT id, status, rolled_back FROM stocktake_batches WHERE id = ?", (batch_id,)
    ).fetchone()
    if batch is None:
        flash("盘点批次不存在")
        return redirect(url_for("stocktake.stocktake_list"))
    if batch["status"] == "approved":
        flash("该批次已审核通过，无需重复操作")
        return redirect(url_for("stocktake.stocktake_list"))
    if int(batch["rolled_back"]) == 1:
        flash("该批次已回滚，无法审核")
        return redirect(url_for("stocktake.stocktake_list"))

    records = db.execute(
        "SELECT item_id, diff FROM stocktakes WHERE batch_id = ?", (batch_id,)
    ).fetchall()
    loss_items, gain_items = [], []
    for record in records:
        diff = parse_qty(record["diff"])
        if diff == 0:
            continue
        item_id = int(record["item_id"])
        (loss_items if diff < 0 else gain_items).append((item_id, diff))

    # Apply stock changes first. Items already touched by a 盘点修正
    # movement for THIS batch (i.e. submitted via /edit before approval)
    # are skipped: the edit path already wrote actual_quantity + diff on
    # the stocktakes row, and approve must not apply diff twice.
    #
    # Scope rule: exact equality (note = ?) on the movement's note string.
    # The note format `修正盘点批次#{batch_id}` is set by submit_edit at
    # the matching line below — if submit_edit ever changes that format,
    # this query must change too. LIKE is NOT acceptable (a substring
    # match like `LIKE '%#X%'` would also catch `#X0`, `#X1`, ..., `#XY`
    # for sibling batch IDs containing `#X`).
    edited_item_ids = {
        r["item_id"] for r in db.execute(
            """SELECT DISTINCT item_id FROM stock_movements
               WHERE action = '盘点修正' AND note = ?""",
            (f"修正盘点批次#{batch_id}",),
        ).fetchall()
    }
    for item_id, diff in loss_items + gain_items:
        if item_id in edited_item_ids:
            continue
        db.execute(
            "UPDATE items SET quantity = quantity + ?, updated_at = ? WHERE id = ?",
            (diff, now(), item_id),
        )

    # 盘亏 → write a synthetic outbound request so summary counts it.
    # Store the new req id in the audit so rollback can find it.
    # Skip items that were already touched by 盘点修正 for this batch:
    # items.quantity is unchanged for those, and a phantom outbound
    # would inflate /summary consumption counts.
    loss_req_ids = []
    for item_id, diff in loss_items:
        if item_id in edited_item_ids:
            continue
        loss_qty = abs(diff)
        cur = db.execute(
            """INSERT INTO outbound_requests
               (item_id, requested_quantity, reason, status, rolled_back, created_at)
               VALUES (?, ?, ?, '出库', 0, ?)""",
            (
                item_id, loss_qty,
                f"盘点审核#{batch_id}盘亏",
                now(),
            ),
        )
        loss_req_ids.append(int(cur.lastrowid))
        db.execute(
            """INSERT INTO stock_movements (item_id, action, delta, note, created_at)
               VALUES (?, '出库', ?, ?, ?)""",
            (item_id, diff, f"盘点审核#{batch_id}盘亏 (req#{loss_req_ids[-1]})", now()),
        )

    # 盘盈 → inventory adjustment only (NOT a consumption).
    # Same idempotency guard: edited items already had a 盘点修正
    # movement; a phantom 库存调整 would double-count the gain.
    for item_id, diff in gain_items:
        if item_id in edited_item_ids:
            continue
        db.execute(
            """INSERT INTO stock_movements (item_id, action, delta, note, created_at)
               VALUES (?, '库存调整', ?, ?, ?)""",
            (item_id, diff, f"盘点审核#{batch_id}盘盈", now()),
        )

    db.execute(
        "UPDATE stocktake_batches SET status = 'approved', loss_req_ids = ? WHERE id = ?",
        (",".join(str(i) for i in loss_req_ids), batch_id),
    )
    db.commit()
    audit("stocktake.approve", "batch", batch_id, {
        "losses": len(loss_items),
        "gains": len(gain_items),
        "loss_req_ids": loss_req_ids,
    })
    msg = f"盘点批次 #{batch_id} 已审核"
    if loss_items:
        msg += f"，{len(loss_items)}项盘亏已计入消耗"
    if gain_items:
        msg += f"，{len(gain_items)}项盘盈已调整为库存"
    flash(msg)
    return redirect(url_for("stocktake.stocktake_list"))
